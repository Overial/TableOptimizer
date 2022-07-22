# Overial
# https://github.com/Overial

# importing necessary modules
import sys
import openpyxl
import pyodbc
import uuid
import os
import datetime
from openpyxl import Workbook
from pathlib import Path

# init today's date
now = datetime.datetime.now()

# init output file path
output_file_path = Path('output_data/optimized_table.xlsx')

# init input files directory
directory = 'input_data/'

# set content first row
content_row = 4

# set id column index
id_col = 2


# connects to server
def connect_to_server(server, database):
    conn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database)
    if conn is not None:
        print('connection succeeded!')
    else:
        print('connection failed')

    return conn.cursor()


# creates table with dates
def update_date_table(cursor):
    for file in os.listdir(directory):
        filename = os.fsdecode(file)
        input_filepath = os.path.join(directory, filename)

        filepaths_raw = cursor.execute('SELECT file_path FROM dbo.FileDates').fetchall()
        filepaths = [filepath[0] for filepath in filepaths_raw]

        if not (input_filepath in filepaths):
            if (filename[0] != '~' and filename[1] != '$') and (filename.endswith('.xlsx')):
                cursor.execute('INSERT INTO dbo.FileDates VALUES (?, ?)', input_filepath, '0')
                cursor.commit()


# loops through filenames and initiates parsing
def init_parsing(cursor):
    filepaths_raw = cursor.execute('SELECT file_path FROM dbo.FileDates WHERE is_processed = 0').fetchall()
    filepaths = [filepath[0] for filepath in filepaths_raw]

    if len(filepaths) == 0:
        print('all files are already processed!')
    else:
        print('starting main loop...')

        for input_file_path in filepaths:
            if output_file_path.is_file():
                parse_excel_file(input_file_path)
            else:
                print('creating optimized table...')
                create_optimized_file(input_file_path)

            cursor.execute('UPDATE dbo.FileDates SET is_processed = 1 WHERE file_path = ?', input_file_path)
            cursor.commit()


# creates blank excel file
def create_excel_file() -> Workbook:
    # create blank Workbook object
    workbook = openpyxl.Workbook()

    return workbook


# creates initial optimized file
def create_optimized_file(input_file_path):
    # get input workbook
    input_workbook = openpyxl.load_workbook(input_file_path)
    # get active input worksheet from input workbook
    input_worksheet = input_workbook.active
    # get active sheet title from input worksheet
    input_worksheet_title = input_worksheet.title
    print(f'active sheet input worksheet title: {input_worksheet_title}')

    # create output workbook
    output_workbook = create_excel_file()
    # get active output worksheet from output workbook
    output_worksheet = output_workbook.active
    # set active sheet title for output worksheet
    output_worksheet.title = input_worksheet_title
    output_worksheet_title = output_worksheet.title
    print(f'active sheet output worksheet title: {output_worksheet_title}')

    # set end data column for storing actual date
    end_date_col = input_worksheet.max_column + 1

    # set deleted flag col
    deleted_flag_col = input_worksheet.max_column + 2

    # set latest flag col
    latest_flag_col = input_worksheet.max_column + 3

    # set current file date
    current_year = int(input_file_path[17:21])
    current_month = int(input_file_path[22:23])
    current_day = int(input_file_path[24:25])
    current_file_date = datetime.datetime(year=current_year, month=current_month, day=current_day)

    for row in range(1, input_worksheet.max_row + 1):
        for col in range(2, input_worksheet.max_column + 1):
            output_worksheet.cell(row, col).value = input_worksheet.cell(row, col).value

            if row == 2 or row == 3:
                output_worksheet.cell(row, end_date_col - 1).value = 'Дата начала'
                output_worksheet.cell(row, end_date_col).value = 'Дата конца'
                output_worksheet.cell(row, deleted_flag_col).value = 'Строка удалена?'
                output_worksheet.cell(row, latest_flag_col).value = 'Актуальная строка?'
            elif row > 3:
                output_worksheet.cell(row, 1).value = str(uuid.uuid4())
                output_worksheet.cell(row, end_date_col).value = current_file_date
                output_worksheet.cell(row, deleted_flag_col).value = 0
                output_worksheet.cell(row, latest_flag_col).value = 1

    output_workbook.save('output_data/optimized_table.xlsx')
    print('initial data read')

    input_workbook.close()
    output_workbook.close()


# parses through input stream
def parse_excel_file(input_file_path):
    print('initiated excel parsing')

    # get input workbook
    input_workbook = openpyxl.load_workbook(input_file_path)
    # get active input worksheet from input workbook
    input_worksheet = input_workbook.active
    # get active sheet title from input worksheet
    input_worksheet_title = input_worksheet.title
    print(f'active sheet input worksheet title: {input_worksheet_title}')

    # set current file date
    current_year = int(input_file_path[17:21])
    current_month = int(input_file_path[21:23])
    current_day = int(input_file_path[23:25])
    current_file_date = datetime.datetime(year=current_year, month=current_month, day=current_day)

    # get output workbook
    output_workbook = openpyxl.load_workbook(output_file_path)
    # get active output worksheet from output workbook
    output_worksheet = output_workbook.active
    # get active sheet title from output worksheet
    output_worksheet_title = output_worksheet.title
    print(f'active sheet output worksheet title: {output_worksheet_title}')

    print()

    # set last content col (input_worksheet)
    last_content_col = input_worksheet.max_column - 1

    # set end data column for storing actual date (output_worksheet)
    end_date_col = output_worksheet.max_column - 2

    # set deleted flag col
    deleted_flag_col = output_worksheet.max_column - 1

    # set latest flag col (output_worksheet)
    latest_flag_col = output_worksheet.max_column

    # iterate through output worksheet
    for output_row in range(content_row, output_worksheet.max_row + 1):
        # get current output cell with id
        output_cell = output_worksheet.cell(output_row, id_col)

        # set row attribute changed flag
        row_attr_changed = False

        # set row found flag
        row_found = False

        # set found row index in case of data update
        found_row_index = 0

        # set new content row
        new_content_row = output_worksheet.max_row + 1

        # check if current row is latest and is not deleted
        if output_worksheet.cell(output_row, latest_flag_col).value == 1:
            # set is row deleted variable
            is_row_deleted = output_worksheet.cell(output_row, deleted_flag_col).value

            # iterate through input worksheet
            for input_row in range(content_row, input_worksheet.max_row + 1):
                # get current input cell with id
                input_cell = input_worksheet.cell(input_row, id_col)

                # found match with latest row
                if output_cell.value == input_cell.value:
                    # set row found flag
                    row_found = True

                    # set found row index
                    found_row_index = input_row

                    # iterate through cols (attributes) in input row
                    for input_col in range(id_col, last_content_col + 1):
                        # get current output attribute cell
                        output_attribute_cell = output_worksheet.cell(output_row, input_col)

                        # get current input attribute cell
                        input_attribute_cell = input_worksheet.cell(input_row, input_col)

                        # float values handling
                        if type(output_attribute_cell.value) == float or type(input_attribute_cell.value) == float:
                            if format(output_cell.value, '.1f') != format(input_cell.value, '.1f'):
                                row_attr_changed = True

                                print(f'[float] found attribute mismatch, row: {input_row}, col: {input_col}')
                                print(f'input: {input_attribute_cell.value}')
                                print(f'output: {output_attribute_cell.value}')

                                break
                        else:
                            if output_attribute_cell.value != input_attribute_cell.value \
                                    and output_attribute_cell.value is not None \
                                    and input_attribute_cell.value is not None:
                                row_attr_changed = True

                                print(f'found attribute mismatch, row: {input_row}, col: {input_col}')
                                print(f'input: {input_attribute_cell.value}')
                                print(f'output: {output_attribute_cell.value}')

                                break

            # update current row latest date
            output_worksheet.cell(output_row, end_date_col).value = current_file_date

            if row_found:
                if row_attr_changed:
                    if is_row_deleted:
                        print(f'row {output_row} returned and changed')
                    else:
                        print(f'row {output_row} changed')

                    # update latest flag cell
                    output_worksheet.cell(output_row, latest_flag_col).value = 0

                    # iterate through attribute cols in input row
                    for new_output_col in range(id_col, last_content_col + 1):
                        # fetch new data from INPUT worksheet
                        new_output_cell = input_worksheet.cell(found_row_index, new_output_col)

                        # write that data to output worksheet
                        output_worksheet.cell(new_content_row, new_output_col).value = new_output_cell.value

                        output_worksheet.cell(new_content_row, 1).value = str(uuid.uuid4())
                        output_worksheet.cell(new_content_row, end_date_col - 1).value = current_file_date
                        output_worksheet.cell(new_content_row, end_date_col).value = current_file_date
                        output_worksheet.cell(new_content_row, deleted_flag_col).value = 0
                        output_worksheet.cell(new_content_row, latest_flag_col).value = 1
                else:
                    # update output row end date attr
                    output_worksheet.cell(output_row, end_date_col).value = current_file_date
            else:
                print(f'row {output_row} was deleted')

                # update latest flag cell
                output_worksheet.cell(output_row, latest_flag_col).value = 0

                # iterate through attributes in output row
                for new_output_col in range(id_col, last_content_col + 1):
                    # fetch old data from output worksheet
                    new_output_cell = output_worksheet.cell(output_row, new_output_col)

                    # write that data to output worksheet
                    output_worksheet.cell(new_content_row, new_output_col).value = new_output_cell.value

                output_worksheet.cell(new_content_row, 1).value = str(uuid.uuid4())
                output_worksheet.cell(new_content_row, end_date_col - 1).value = current_file_date
                output_worksheet.cell(new_content_row, end_date_col).value = current_file_date
                # set deleted flag to 1
                output_worksheet.cell(new_content_row, deleted_flag_col).value = 1
                output_worksheet.cell(new_content_row, latest_flag_col).value = 1

    output_workbook.save(output_file_path)

    input_workbook.close()
    output_workbook.close()


# main function
def main() -> int:
    print('initiated table optimization' + str(now))

    server = '.\\SQLEXPRESS'
    database = 'Test'
    cursor = connect_to_server(server, database)

    # update date table
    update_date_table(cursor)

    # start main loop
    init_parsing(cursor)

    # close connection
    cursor.connection.close()

    return 0


if __name__ == '__main__':
    sys.exit(main())
